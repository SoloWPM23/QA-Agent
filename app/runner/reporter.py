"""Report renderers: turn a run summary into Markdown, JSON, or Excel."""

from __future__ import annotations

import json
from io import BytesIO
from typing import Any

from app.runner.base import register_reporter


@register_reporter("markdown")
class MarkdownReporter:
    """Render a run summary as a Markdown report."""

    def render(self, summary: dict) -> str:
        lines: list[str] = ["# Test Suite Execution Report", ""]
        if summary.get("base_url"):
            lines.append(f"**Base URL:** `{summary['base_url']}`")
            lines.append("")
        lines.append(
            f"**Total:** {summary.get('total', 0)} | "
            f"**PASS:** {summary.get('passed', 0)} | "
            f"**FAIL:** {summary.get('failed', 0)} | "
            f"**SKIPPED:** {summary.get('skipped', 0)}"
        )
        lines.append("")
        lines.append("| Test Case | Status | Reason | Explanation |")
        lines.append("|-----------|--------|--------|-------------|")
        for verdict in summary.get("verdicts", []):
            cid = verdict.get("case_id", "(no-id)")
            status = verdict.get("status", "?")
            reason = (verdict.get("reason") or "").replace("|", "\\|")
            explanation = (verdict.get("explanation") or "").replace("|", "\\|")
            lines.append(f"| {cid} | {status} | {reason} | {explanation} |")
        lines.append("")
        return "\n".join(lines)


@register_reporter("json")
class JsonReporter:
    """Render a run summary as pretty-printed JSON."""

    def render(self, summary: dict) -> str:
        return json.dumps(summary, indent=2, ensure_ascii=False)


@register_reporter("excel")
class ExcelReporter:
    """Render a run summary as an Excel workbook (bytes)."""

    def render(self, summary: dict) -> bytes:
        return render_excel(summary)


def render_reports(summary: dict, formats: list[str]) -> dict[str, str | bytes]:
    """Produce rendered reports for each requested format name."""
    from app.runner.base import get_reporter

    out: dict[str, str | bytes] = {}
    for fmt in formats:
        reporter = get_reporter(fmt)()
        out[fmt] = reporter.render(summary)
    return out


def render_excel(summary: dict) -> bytes:
    """Build an .xlsx workbook from a run summary and return its bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    summary_sheet = wb.active
    if summary_sheet is None:
        summary_sheet = wb.create_sheet("Summary")
    else:
        summary_sheet.title = "Summary"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E3F2FD")

    summary_sheet.append(["Metric", "Value"])
    for cell in summary_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    summary_sheet.append(["Base URL", summary.get("base_url", "")])
    summary_sheet.append(["Total", summary.get("total", 0)])
    summary_sheet.append(["PASS", summary.get("passed", 0)])
    summary_sheet.append(["FAIL", summary.get("failed", 0)])
    summary_sheet.append(["SKIPPED", summary.get("skipped", 0)])
    summary_sheet.column_dimensions["A"].width = 18
    summary_sheet.column_dimensions["B"].width = 50

    detail_sheet = wb.create_sheet("Details")
    headers = [
        "Case ID",
        "Title",
        "Status",
        "Method",
        "Path",
        "Reason",
        "Explanation",
        "Request Body",
        "Response Status",
        "Response Body",
        "Assertions",
    ]
    detail_sheet.append(headers)
    for cell in detail_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    for verdict in summary.get("verdicts", []):
        detail = verdict.get("detail", {})
        assertions = verdict.get("assertions", [])
        assertion_text = "\n".join(
            f"{a.get('name')} - passed={a.get('passed')} - {a.get('details')}" for a in assertions
        )
        detail_sheet.append(
            [
                verdict.get("case_id", ""),
                detail.get("title", ""),
                verdict.get("status", ""),
                detail.get("method", ""),
                detail.get("path", ""),
                verdict.get("reason", ""),
                verdict.get("explanation", ""),
                _as_text(detail.get("request_body")),
                detail.get("response_status", ""),
                _as_text(detail.get("response_body")),
                assertion_text,
            ]
        )

    for col in detail_sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            length = len(str(cell.value))
            max_length = max(max_length, length)
        detail_sheet.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 80)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False)


def _write_file(path: str, content: str, encoding: str = "utf-8") -> None:
    with open(path, "w", encoding=encoding) as fh:
        fh.write(content)


def persist_reports(summary: dict, format_paths: dict[str, str]) -> list[str]:
    """Render a summary and write each format to its configured path."""
    written: list[str] = []
    for fmt, path in format_paths.items():
        rendered = render_reports(summary, [fmt])[fmt]
        if isinstance(rendered, bytes):
            with open(path, "wb") as fh:
                fh.write(rendered)
        else:
            _write_file(path, rendered)
        written.append(path)
    return written
